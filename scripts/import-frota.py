# Importa "Controle de Frota 1.xlsx" para as tabelas frota_* do Supabase.
# Uso: python scripts/import-frota.py  (requer .env.local com SUPABASE_SERVICE_ROLE_KEY)
import json
import os
import re
import sys
import urllib.request
from datetime import date, datetime, time

import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "Controle de Frota 1.xlsx")

env = {}
with open(os.path.join(ROOT, ".env.local"), encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()

URL = env["NEXT_PUBLIC_SUPABASE_URL"]
KEY = env["SUPABASE_SERVICE_ROLE_KEY"]

NULLISH = {"", "N/I", "N/A", "N/D", "?", "-", "N/ I"}


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return None if v.upper() in NULLISH else v
    return v


def as_text(v):
    v = clean(v)
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def as_date(v):
    v = clean(v)
    if isinstance(v, datetime):
        v = v.date()
    if isinstance(v, date):
        return None if v.year < 1990 else v.isoformat()
    if isinstance(v, str):
        m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", v)
        if m:
            return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return None


def as_num(v):
    v = clean(v)
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, str):
        s = v.replace("R$", "").replace(".", "").replace(",", ".").strip()
        try:
            return round(float(s), 2)
        except ValueError:
            return None
    return None


def as_int(v):
    n = as_num(v)
    return int(n) if n is not None else None


PLACA_RE = re.compile(r"^[A-Z]{3}[0-9][A-Z0-9][0-9]{2}$")


def as_placa(v):
    v = as_text(v)
    if v is None:
        return None
    v = v.upper().replace(" ", "").replace("-", "")
    return v if PLACA_RE.match(v) else None


def post(table, rows):
    if not rows:
        return 0
    # PostgREST exige chaves uniformes em inserts em lote
    keys = {k for r in rows for k in r}
    rows = [{k: r.get(k) for k in keys} for r in rows]
    req = urllib.request.Request(
        f"{URL}/rest/v1/{table}",
        data=json.dumps(rows).encode(),
        headers={
            "apikey": KEY,
            "Authorization": f"Bearer {KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req) as resp:
            assert resp.status in (200, 201), resp.status
    except urllib.error.HTTPError as e:
        print(f"ERRO {table}: {e.code} {e.read().decode(errors='replace')[:500]}")
        raise
    return len(rows)


def get_ids(table, col="placa"):
    req = urllib.request.Request(
        f"{URL}/rest/v1/{table}?select=id,{col}",
        headers={"apikey": KEY, "Authorization": f"Bearer {KEY}"},
    )
    with urllib.request.urlopen(req) as resp:
        return {r[col]: r["id"] for r in json.loads(resp.read()) if r.get(col)}


wb = openpyxl.load_workbook(XLSX, data_only=True)

# ---------- fornecedores ----------
ws = wb["Locadoras-Rastreadores-Veloe"]
fornecedores = []
for row in ws.iter_rows(min_row=2, values_only=True):
    empresa = as_text(row[1])
    if not empresa or empresa.isdigit():
        continue
    fornecedores.append({
        "status": as_text(row[0]),
        "empresa": empresa,
        "atendimento": as_text(row[2]),
        "servicos": as_text(row[3]),
        "contato": as_text(row[4]),
        "telefone": as_text(row[5]),
        "whatsapp": as_text(row[6]),
        "site_email": as_text(row[9]),  # login/senha (cols 7-8) não são importados
    })
print("fornecedores:", post("frota_fornecedores", fornecedores))

# ---------- prestadores ----------
ws = wb["Prestadores"]
prestadores = []
for row in ws.iter_rows(min_row=2, values_only=True):
    nome = as_text(row[2])
    if not nome or nome.isdigit():
        continue
    prestadores.append({
        "cidade": as_text(row[0]),
        "classificacao": as_text(row[1]),
        "nome": nome,
        "link_maps": as_text(row[3]),
        "telefone": as_text(row[4]),
        "contato": as_text(row[5]),
        "motivo_classificacao": as_text(row[6]),
    })
print("prestadores:", post("frota_prestadores", prestadores))

# ---------- veículos (aba Controle, cabeçalho na linha 2) ----------
ws = wb["Controle"]
veiculos = []
controle_rows = []
for row in ws.iter_rows(min_row=3, values_only=True):
    placa = as_placa(row[9])
    if not placa:
        continue
    controle_rows.append((placa, row))
    veiculos.append({
        "placa": placa,
        "renavam": as_text(row[19]),
        "crv": as_text(row[21]),
        "uf": as_text(row[22]),
        "marca": as_text(row[7]),
        "modelo": as_text(row[8]),
        "tipo": as_text(row[6]),
        "ano_fabricacao": as_text(row[23]),
        "exercicio_crlv": as_text(row[24]),
        "combustivel": as_text(row[25]),
        "chave_reserva": as_text(row[43]),
        "status": as_text(row[1]) or "INATIVO",
        "acao": as_text(row[2]),
        "aplicacao_devolucao": as_text(row[60]),
        "data_aplicacao": as_date(row[61]),
        "gestor": as_text(row[0]),
        "local_trabalho": as_text(row[3]),
        "centro_custo": as_text(row[4]),
        "ut_atual": as_text(row[5]),
        "condutor_nome": as_text(row[11]),
        "condutor_re": as_text(row[10]),
        "telefone": as_text(row[14]),
        "score_dirigibilidade": as_text(row[15]),
        "validade_credenciamento": as_date(row[16]),
        "propriedade": as_text(row[13]),
        "modalidade": as_text(row[17]),
        "tipo_contrato": as_text(row[42]),
        "valor_locacao": as_num(row[18]),
        "valor_aporte": as_num(row[12]),
        "cnpj_proprietario": as_text(row[20]),
        "rastreador": as_text(row[36]),
        "num_requisicao": as_text(row[39]),
        "data_requisicao": as_date(row[40]),
        "status_rv": as_text(row[41]),
        "chamado": as_text(row[37]),
        "conteudo": as_text(row[38]),
    })
# dedup por placa (mantém a primeira ocorrência)
seen = set()
veiculos = [v for v in veiculos if not (v["placa"] in seen or seen.add(v["placa"]))]
print("veiculos:", post("frota_veiculos", veiculos))
placa_id = get_ids("frota_veiculos")

# ---------- manutenções (aba Manutenção) ----------
ws = wb["Manutenção"]
manutencoes = []
mkeys = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    placa = as_placa(row[3])
    tipo = as_text(row[2])
    if not placa or tipo not in ("PREVENTIVA", "CORRETIVA", "SINISTRO"):
        continue
    m = {
        "veiculo_id": placa_id.get(placa),
        "placa": placa,
        "tipo": tipo,
        "situacao": as_text(row[1]),
        "km_atual": as_int(row[7]),
        "previsao_proxima": as_date(row[8]),
        "km_proxima_revisao": as_int(row[9]),
        "data_atendimento": as_date(row[10]),
        "data_parada": as_date(row[11]),
        "hora": as_text(row[12]),
        "local_oficina": as_text(row[13]),
        "descricao_servico": as_text(row[14]),
        "protocolo": as_text(row[15]),
        "diretor": as_text(row[16]),
        "coligada": as_text(row[17]),
    }
    mkeys.add((placa, m["tipo"], m["data_parada"], m["descricao_servico"]))
    manutencoes.append(m)

# última preventiva/corretiva embutidas na aba Controle (se não estiverem no histórico)
for placa, row in controle_rows:
    prev = {
        "veiculo_id": placa_id.get(placa),
        "placa": placa,
        "tipo": "PREVENTIVA",
        "situacao": "FINALIZADO",
        "data_atendimento": as_date(row[44]),
        "data_parada": as_date(row[45]),
        "hora": as_text(row[46]),
        "km_atual": as_int(row[47]),
        "previsao_proxima": as_date(row[48]),
        "km_proxima_revisao": as_int(row[49]),
        "local_oficina": as_text(row[50]),
        "descricao_servico": as_text(row[51]),
        "protocolo": as_text(row[52]),
    }
    corr = {
        "veiculo_id": placa_id.get(placa),
        "placa": placa,
        "tipo": "CORRETIVA",
        "situacao": "FINALIZADO",
        "data_atendimento": as_date(row[53]),
        "data_parada": as_date(row[54]),
        "hora": as_text(row[55]),
        "km_atual": as_int(row[56]),
        "local_oficina": as_text(row[57]),
        "descricao_servico": as_text(row[58]),
        "protocolo": as_text(row[59]),
    }
    for m in (prev, corr):
        if not (m["data_parada"] or m["descricao_servico"]):
            continue
        key = (placa, m["tipo"], m["data_parada"], m["descricao_servico"])
        if key not in mkeys:
            mkeys.add(key)
            manutencoes.append(m)
print("manutencoes:", post("frota_manutencoes", manutencoes))

# ---------- cartões (Alelo estoque + aba Controle) ----------
CARD_RE = re.compile(r"^[0-9 *•?]{10,}$")
cartoes = {}
ws = wb["Alelo (estoque)"]
for row in ws.iter_rows(min_row=2, values_only=True):
    numero = as_text(row[1])
    if not numero or not CARD_RE.match(numero):
        continue
    cartoes[numero] = {
        "numero": numero,
        "status": as_text(row[0]),
        "tipo": as_text(row[3]) or "ESTOQUE",
        "ultima_placa": as_placa(row[4]),
        "ultimo_condutor": as_text(row[5]),
    }
for placa, row in controle_rows:
    numero = as_text(row[28])
    if not numero or not CARD_RE.match(numero) or numero in cartoes:
        continue
    cartoes[numero] = {
        "numero": numero,
        "status": as_text(row[26]),
        "tipo": as_text(row[27]) or "VEÍCULO",
        "veiculo_id": placa_id.get(placa),
        "limite_anterior": as_num(row[30]),
        "limite_atual": as_num(row[32]),
        "saldo_atual": as_num(row[33]),
        "ultima_placa": placa,
    }
print("cartoes:", post("frota_cartoes", list(cartoes.values())))

# ---------- tags Veloe (estoque + aba Controle) ----------
tags = {}
ws = wb["Veloe (estoque)"]
for row in ws.iter_rows(min_row=2, values_only=True):
    numero = as_text(row[1])
    if not numero or not numero.isdigit():
        continue
    tags[numero] = {"numero": numero, "status": as_text(row[0]), "marca": as_text(row[2])}
for placa, row in controle_rows:
    numero = as_text(row[35])
    if numero:
        numero = numero.strip()
    if not numero or not numero.isdigit() or numero in tags:
        continue
    tags[numero] = {
        "numero": numero,
        "status": as_text(row[34]),
        "marca": "VELOE",
        "veiculo_id": placa_id.get(placa),
    }
print("tags:", post("frota_tags", list(tags.values())))

print("OK")
